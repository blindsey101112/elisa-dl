import os
import sys
import shutil
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import leastsq
from scipy.stats import variation
import pdfkit
import datetime

'''
Credit to https://people.duke.edu/~ccc14/pcfb/analysis.html for the code to fit 
the 4 parameter logistic regression for the standard curve
'''

def get_ods(file):
    wb = load_workbook(file)
    od_ws = wb["Photometric1"]
    ods={}
    std_curve1 = {}
    for cell in od_ws["B23":"M23"][0]:
        std_curve1[cell.coordinate] = cell.value
    ods["std_curve1"] = std_curve1
    std_curve2 = {}
    for cell in od_ws["B24":"M24"][0]:
        std_curve2[cell.coordinate] = cell.value
    ods["std_curve2"] = std_curve2
    pos = {}
    for cell in od_ws["F22":"G22"][0]:
        pos[cell.coordinate] = cell.value
    ods["pos"] = pos
    blk = {}
    for cell in od_ws["J22":"M22"][0]:
        blk[cell.coordinate] = cell.value
    ods["blk"] =  blk
    neg = {}
    for cell in od_ws["H22":"I22"][0]:
        neg[cell.coordinate] = cell.value
    ods["neg"] = neg
    sample01 = {}
    for cell in od_ws["B17":"C17"][0]:
        sample01[cell.coordinate] = cell.value
    ods["sample01"] = sample01
    sample02 = {}
    for cell in od_ws["D17":"E17"][0]:
        sample02[cell.coordinate] = cell.value
    ods["sample02"] = sample02
    sample03 = {}
    for cell in od_ws["F17":"G17"][0]:
        sample03[cell.coordinate] = cell.value
    ods["sample03"] = sample03
    sample04 = {}
    for cell in od_ws["H17":"I17"][0]:
        sample04[cell.coordinate] = cell.value
    ods["sample04"] = sample04
    sample05 = {}
    for cell in od_ws["J17":"K17"][0]:
        sample05[cell.coordinate] = cell.value
    ods["sample05"] = sample05
    sample06 = {}
    for cell in od_ws["L17":"M17"][0]:
        sample06[cell.coordinate] = cell.value
    ods["sample06"] = sample06
    sample07 = {}
    for cell in od_ws["B18":"C18"][0]:
        sample07[cell.coordinate] = cell.value
    ods["sample07"] = sample07
    sample08 = {}
    for cell in od_ws["D18":"E18"][0]:
        sample08[cell.coordinate] = cell.value
    ods["sample08"] = sample08
    sample09 = {}
    for cell in od_ws["F18":"G18"][0]:
        sample09[cell.coordinate] = cell.value
    ods["sample09"] = sample09
    sample10 = {}
    for cell in od_ws["H18":"I18"][0]:
        sample10[cell.coordinate] = cell.value
    ods["sample10"] = sample10
    sample11 = {}
    for cell in od_ws["J18":"K18"][0]:
        sample11[cell.coordinate] = cell.value
    ods["sample11"] = sample11
    sample12 = {}
    for cell in od_ws["L18":"M18"][0]:
        sample12[cell.coordinate] = cell.value
    ods["sample12"] = sample12
    sample13 = {}
    for cell in od_ws["B19":"C19"][0]:
        sample13[cell.coordinate] = cell.value
    ods["sample13"] = sample13
    sample14 = {}
    for cell in od_ws["D19":"E19"][0]:
        sample14[cell.coordinate] = cell.value
    ods["sample14"] = sample14
    sample15 = {}
    for cell in od_ws["F19":"G19"][0]:
        sample15[cell.coordinate] = cell.value
    ods["sample15"] = sample15
    sample16 = {}
    for cell in od_ws["H19":"I19"][0]:
        sample16[cell.coordinate] = cell.value
    ods["sample16"] = sample16
    sample17 = {}
    for cell in od_ws["J19":"K19"][0]:
        sample17[cell.coordinate] = cell.value
    ods["sample17"] = sample17
    sample18 = {}
    for cell in od_ws["L19":"M19"][0]:
        sample18[cell.coordinate] = cell.value
    ods["sample18"] = sample18
    sample19 = {}
    for cell in od_ws["B20":"C20"][0]:
        sample19[cell.coordinate] = cell.value
    ods["sample19"] = sample19
    sample20 = {}
    for cell in od_ws["D20":"E20"][0]:
        sample20[cell.coordinate] = cell.value
    ods["sample20"] = sample20
    sample21 = {}
    for cell in od_ws["F20":"G20"][0]:
        sample21[cell.coordinate] = cell.value
    ods["sample21"] = sample21
    sample22 = {}
    for cell in od_ws["H20":"I20"][0]:
        sample22[cell.coordinate] = cell.value
    ods["sample22"] = sample22
    sample23 = {}
    for cell in od_ws["J20":"K20"][0]:
        sample23[cell.coordinate] = cell.value
    ods["sample23"] = sample23
    sample24 = {}
    for cell in od_ws["L20":"M20"][0]:
        sample24[cell.coordinate] = cell.value
    ods["sample24"] = sample24
    sample25 = {}
    for cell in od_ws["B21":"C21"][0]:
        sample25[cell.coordinate] = cell.value
    ods["sample25"] = sample25
    sample26 = {}
    for cell in od_ws["D21":"E21"][0]:
        sample26[cell.coordinate] = cell.value
    ods["sample26"] = sample26
    sample27 = {}
    for cell in od_ws["F21":"G21"][0]:
        sample27[cell.coordinate] = cell.value
    ods["sample27"] = sample27
    sample28 = {}
    for cell in od_ws["H21":"I21"][0]:
        sample28[cell.coordinate] = cell.value
    ods["sample28"] = sample28
    sample29 = {}
    for cell in od_ws["J21":"K21"][0]:
        sample29[cell.coordinate] = cell.value
    ods["sample29"] = sample29
    sample30 = {}
    for cell in od_ws["L21":"M21"][0]:
        sample30[cell.coordinate] = cell.value
    ods["sample30"] = sample30
    sample31 = {}
    for cell in od_ws["B22":"C22"][0]:
        sample31[cell.coordinate] = cell.value
    ods["sample31"] = sample31
    sample32 = {}
    for cell in od_ws["D22":"E22"][0]:
        sample32[cell.coordinate] = cell.value
    ods["sample32"] = sample32
    return ods

def get_samples(file):
    wb = load_workbook(file)
    sample_ws = wb["PlatePlan"]
    sample_dilution = {
               "sample01" : sample_ws["B17"].value,
               "sample02" : sample_ws["D17"].value,
               "sample03" : sample_ws["F17"].value,
               "sample04" : sample_ws["H17"].value,
               "sample05" : sample_ws["J17"].value,
               "sample06": sample_ws["L17"].value,
               "sample07": sample_ws["B18"].value,
               "sample08": sample_ws["D18"].value,
               "sample09": sample_ws["F18"].value,
               "sample10": sample_ws["H18"].value,
               "sample11": sample_ws["J18"].value,
               "sample12": sample_ws["L18"].value,
               "sample13": sample_ws["B19"].value,
               "sample14": sample_ws["D19"].value,
               "sample15": sample_ws["F19"].value,
               "sample16": sample_ws["H19"].value,
               "sample17": sample_ws["J19"].value,
               "sample18": sample_ws["L19"].value,
               "sample19": sample_ws["B20"].value,
               "sample20": sample_ws["D20"].value,
               "sample21": sample_ws["F20"].value,
               "sample22": sample_ws["H20"].value,
               "sample23": sample_ws["J20"].value,
               "sample24": sample_ws["L20"].value,
               "sample25": sample_ws["B21"].value,
               "sample26": sample_ws["D21"].value,
               "sample27": sample_ws["F21"].value,
               "sample28": sample_ws["H21"].value,
               "sample29": sample_ws["J21"].value,
               "sample30": sample_ws["L21"].value,
               "sample31": sample_ws["B22"].value,
               "sample32": sample_ws["D22"].value
               }
    return sample_dilution

def logistic4(x, A, B, C, D):
    """4PL lgoistic equation."""
    asym_dif = A-D
    con_inflec = x/C
    gradiant_power = np.sign(con_inflec) * (np.abs(con_inflec)) ** B
    equation = (asym_dif/(1.0 + gradiant_power) + D)
    return equation

def residuals(p, y, x):
    """Deviations of data from fitted 4PL curve"""
    A,B,C,D = p
    err = y-logistic4(x, A, B, C, D)
    return err

def peval(x, p):
    """Evaluated value at x with current parameters."""
    A,B,C,D = p
    return logistic4(x, A, B, C, D)

def get_conc(y, p):
    A,B,C,D = p
    step1 = ((A-D)/(y-D)) - 1
    step2 = np.sign(step1) * (np.abs(step1)) ** (1/B)
    concentration = step2 * C
    return concentration

std_concs = [1000, 571.4285714, 326.5306122, 186.5889213, 106.6222407, 60.9269947,
             34.81542555, 19.89452888, 11.36830222, 6.496172697, 3.712098684, 2.121199248]

if __name__ == "__main__":
    plate_id = sys.argv[1]
    plateplan_file = plate_id + "-pplan.xlsx"
    platereader_file = plate_id + "-preader.xlsx"
    ods = get_ods(platereader_file)
    sample_dilution = get_samples(plateplan_file)

    print("Found plateplan file: %s" % plateplan_file)
    print("Found plate reader file: %s" % platereader_file)

    np.set_printoptions(suppress=True)

### Fit standard curve using 4 parameter logistic regression ###
    print("Fitting standard curve")
    x = np.asarray(std_concs)

    std_curve1_ods = np.asarray(list(ods["std_curve1"].values()))
    std_curve2_ods = np.asarray(list(ods["std_curve2"].values()))

    y = (std_curve1_ods + std_curve2_ods) / 2.0

    # Initial guess for parameters
    p0 = [0, 1, 1, 1]

    # Fit equation using least squares optimization
    plsq = leastsq(residuals, p0, args=(y, x))

    # Plot results
    plt.plot(x, peval(x, plsq[0]))
    plt.plot(x, std_curve1_ods, '.', color='orange')
    plt.plot(x, std_curve2_ods, '.', color='orange')

    plt.xscale("log", basex=10)
    plt.title("Standard curve")
    plt.xlabel("Unit of standard")
    plt.ylabel("OD")

    fig_name = plate_id + ".png"
    fig_path = os.path.join("figs", fig_name)

    plt.savefig(fig_path)


### sample concentrations ###
    print("Calculating concentrations")
    sample_means = {}
    sample_cv = {}
    sample_concs = {}

    for sample in ods.keys():
        if "sample" in sample:
            sample_ods = np.asarray(list(ods[sample].values()))
            mean = sum(sample_ods)/len(sample_ods)
            sample_concs[sample] = round(get_conc(mean, plsq[0]), 6)
            cv = variation(sample_ods, axis = 0)
            if sample_concs[sample] < std_concs[-1]:
                sample_concs[sample] = "BelowCurve"
            elif sample_concs[sample] > std_concs[0]:
                sample_concs[sample] = "AboveCurve"
            sample_means[sample] = round(mean, 3)
            sample_cv[sample] = round(cv, 2)



### Output to pdf file ###
    print("Generating html file")
    now = datetime.datetime.now()
    date = "%s-%s-%s" % (now.day, now.strftime("%b"), now.year)

    html_template = """
    <html>
    <body>
    
    <style>
    .centre {
        text-align: center;
    }
        
        table {
      font-family: arial, sans-serif;
      border-collapse: collapse;
    }
    
    td, th {
      border: 1px solid #dddddd;
      text-align: left;
      padding: 8px;
    }
    
    tr:nth-child(1) {
      background-color: #dddddd;
    }
    </style>
    
    <h1> Plate Report - %s</h1>
    
    <p> Report generated on %s<p>
    
    <p class="centre"><img src="%s" alt="Standard curve" width="450" height="350"/></p>
    
    <table>
      <tr>
        <th>SampleID</th>
        <th>OD</th>
        <th>CV</th>
        <th>Ab-Units</th>
      </tr>
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>      
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>      
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>      
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>      
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>      
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>      
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>      
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
      </tr>      
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>          
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>          
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>        
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>  
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>  
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>  
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>  
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>  
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>  
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>  
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>  
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>  
      <tr>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>
        <td>%s</td>        
      </tr>  
      
    </table>
    
    </body>
    </html>
    
    """

    html = html_template % (plate_id,
                            date,
                            fig_path,

                            sample_dilution["sample01"].split("-")[0],
                                sample_means["sample01"],
                                sample_cv["sample01"],
                                sample_concs["sample01"],

                            sample_dilution["sample02"].split("-")[0],
                                sample_means["sample02"],
                                sample_cv["sample02"],
                                sample_concs["sample02"],

                            sample_dilution["sample03"].split("-")[0],
                                sample_means["sample03"],
                                sample_cv["sample03"],
                                sample_concs["sample03"],

                            sample_dilution["sample04"].split("-")[0],
                                sample_means["sample04"],
                                sample_cv["sample04"],
                                sample_concs["sample04"],

                            sample_dilution["sample05"].split("-")[0],
                                sample_means["sample05"],
                                sample_cv["sample05"],
                                sample_concs["sample05"],

                            sample_dilution["sample06"].split("-")[0],
                                sample_means["sample06"],
                                sample_cv["sample06"],
                                sample_concs["sample06"],

                            sample_dilution["sample07"].split("-")[0],
                                sample_means["sample07"],
                                sample_cv["sample07"],
                                sample_concs["sample07"],

                            sample_dilution["sample08"].split("-")[0],
                                sample_means["sample08"],
                                sample_cv["sample08"],
                                sample_concs["sample08"],

                            sample_dilution["sample09"].split("-")[0],
                                sample_means["sample09"],
                                sample_cv["sample09"],
                                sample_concs["sample09"],


                            sample_dilution["sample10"].split("-")[0],
                                sample_means["sample10"],
                                sample_cv["sample10"],
                                sample_concs["sample10"],

                            sample_dilution["sample11"].split("-")[0],
                                sample_means["sample11"],
                                sample_cv["sample11"],
                                sample_concs["sample11"],

                            sample_dilution["sample12"].split("-")[0],
                                sample_means["sample12"],
                                sample_cv["sample12"],
                                sample_concs["sample12"],

                            sample_dilution["sample13"].split("-")[0],
                                sample_means["sample13"],
                                sample_cv["sample13"],
                                sample_concs["sample13"],

                            sample_dilution["sample14"].split("-")[0],
                                sample_means["sample14"],
                                sample_cv["sample14"],
                                sample_concs["sample14"],

                            sample_dilution["sample15"].split("-")[0],
                                sample_means["sample15"],
                                sample_cv["sample15"],
                                sample_concs["sample15"],

                            sample_dilution["sample16"].split("-")[0],
                                sample_means["sample16"],
                                sample_cv["sample16"],
                                sample_concs["sample16"],

                            sample_dilution["sample17"].split("-")[0],
                                sample_means["sample17"],
                                sample_cv["sample17"],
                                sample_concs["sample17"],

                            sample_dilution["sample18"].split("-")[0],
                                sample_means["sample18"],
                                sample_cv["sample18"],
                                sample_concs["sample18"],

                            sample_dilution["sample19"].split("-")[0],
                                sample_means["sample19"],
                                sample_cv["sample19"],
                                sample_concs["sample19"],


                            sample_dilution["sample20"].split("-")[0],
                                sample_means["sample20"],
                                sample_cv["sample20"],
                                sample_concs["sample20"],

                            sample_dilution["sample21"].split("-")[0],
                                sample_means["sample21"],
                                sample_cv["sample21"],
                                sample_concs["sample21"],

                            sample_dilution["sample22"].split("-")[0],
                                sample_means["sample22"],
                                sample_cv["sample22"],
                                sample_concs["sample22"],

                            sample_dilution["sample23"].split("-")[0],
                                sample_means["sample23"],
                                sample_cv["sample23"],
                                sample_concs["sample23"],

                            sample_dilution["sample24"].split("-")[0],
                                sample_means["sample24"],
                                sample_cv["sample24"],
                                sample_concs["sample24"],

                            sample_dilution["sample25"].split("-")[0],
                                sample_means["sample25"],
                                sample_cv["sample25"],
                                sample_concs["sample25"],

                            sample_dilution["sample26"].split("-")[0],
                                sample_means["sample26"],
                                sample_cv["sample26"],
                                sample_concs["sample26"],

                            sample_dilution["sample27"].split("-")[0],
                                sample_means["sample27"],
                                sample_cv["sample27"],
                                sample_concs["sample27"],

                            sample_dilution["sample28"].split("-")[0],
                                sample_means["sample28"],
                                sample_cv["sample28"],
                                sample_concs["sample28"],

                            sample_dilution["sample29"].split("-")[0],
                                sample_means["sample29"],
                                sample_cv["sample29"],
                                sample_concs["sample29"],

                            sample_dilution["sample30"].split("-")[0],
                                sample_means["sample30"],
                                sample_cv["sample30"],
                                sample_concs["sample30"],

                            sample_dilution["sample31"].split("-")[0],
                                sample_means["sample31"],
                                sample_cv["sample31"],
                                sample_concs["sample31"],

                            sample_dilution["sample32"].split("-")[0],
                                sample_means["sample32"],
                                sample_cv["sample32"],
                                sample_concs["sample32"]

                            )

    html_file = plate_id + ".html"
    pdf_file = plate_id + ".pdf"

    with open(plate_id + ".html", 'w') as htmlfile:
        htmlfile.write(html)


    print("Converting html to pdf...")
    pdfkit.from_file(html_file, pdf_file)

    shutil.move(html_file, os.path.join("html_reports", html_file))

